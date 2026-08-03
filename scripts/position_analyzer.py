import asyncio
from logging import basicConfig, INFO, getLogger
from typing import List, Optional, Tuple

import pandas as pd
import argparse
import textwrap
import traceback

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.base_driver import BaseDriver
from core.common import SecurityDescriptor, OptionInfo, BarSize, RequestedInfoType, CoreException
from core.options_data import OptionData, OptionDataException
from core.option_data_manager import OptionDataManager
from core.ib.ib_driver import IBDriver
from core.schwab.schwab_driver import SchwabDriver
from core.utils import current_datetime, calculate_expected_move, get_datetime

r"""
Utility for analyzing a set of open options positions. For each option leg it reports current
price and the Greeks (delta, theta, gamma, vega) alongside the trade details recorded in a CSV,
plus the contracts actually held (Quantity + Quantity Out) and any realized P/L on the closed
portion. A final aggregate row summarizes the filtered legs (net position Greeks, net value,
unrealized P/L, and realized P/L), computed over the contracts actually held. Finally, it prints
the expected next-day (1 std dev) move for each underlying, from the broker's implied volatility.

Setup and usage:
------------------------
d:
cd CodingProjects\Python\TWS2025
conda activate options_2025_1
python -m scripts.position_analyzer --help    # full instruction manual with examples
"""

CLIENT_ID = 17
# Standard US equity-option contract multiplier: 1 contract controls 100 shares.
CONTRACT_MULTIPLIER = 100

# CSV column names (current_positions.csv format:
# Position #,Date In,Position Type,Symbol,Quantity,Trade Price,Date Out,Quantity Out,Exit Price)
CSV_POSITION_NUM = "Position #"
CSV_DATE_IN = "Date In"
CSV_POSITION_TYPE = "Position Type"
CSV_SYMBOL = "Symbol"
CSV_QUANTITY = "Quantity"
CSV_TRADE_PRICE = "Trade Price"
CSV_DATE_OUT = "Date Out"
CSV_QUANTITY_OUT = "Quantity Out"
CSV_EXIT_PRICE = "Exit Price"

# Every column the tool requires to be present in the positions CSV (validated up front).
REQUIRED_COLUMNS = [
    CSV_POSITION_NUM,
    CSV_DATE_IN,
    CSV_POSITION_TYPE,
    CSV_SYMBOL,
    CSV_QUANTITY,
    CSV_TRADE_PRICE,
    CSV_DATE_OUT,
    CSV_QUANTITY_OUT,
    CSV_EXIT_PRICE,
]

# Maps the short --position-type argument to the full name stored in the CSV's "Position Type" column.
POSITION_TYPE_MAP = {
    "IC": "Iron Condor",
    "CS": "Credit Spread",
    "DS": "Debit Spread",
    "L": "Naked Long",
    "S": "Naked Short",
    "LC": "Long Call",
    "CSP": "Cash Secured Put",
    "CAL": "Calendar",
    "DCAL": "Double Calendar",
    "TCAL": "Triple Calendar",
    "DIAG": "Diagonal",
    "DDIAG": "Double Diagonal",
}

# Output column names, in display order
COL_CONTRACT = "Contract"
COL_POSITION_NUM = "Pos #"
COL_POSITION_TYPE = "Pos Type"
COL_QUANTITY = "Qty"
COL_HELD = "Held"
COL_TRADE_PRICE = "Trade Price"
COL_CUR_EXIT_PRICE = "Cur/Exit Price"
COL_REALIZED = "Realized"
COL_IV = "IV"
COL_DELTA = "Delta"
COL_THETA = "Theta"
COL_GAMMA = "Gamma"
COL_VEGA = "Vega"

OUTPUT_COLUMNS = [
    COL_CONTRACT,
    COL_POSITION_NUM,
    COL_POSITION_TYPE,
    COL_QUANTITY,
    COL_HELD,
    COL_TRADE_PRICE,
    COL_CUR_EXIT_PRICE,
    COL_REALIZED,
    COL_IV,
    COL_DELTA,
    COL_THETA,
    COL_GAMMA,
    COL_VEGA,
]

# Column names for the --show summary table (one row per whole position)
COL_SHOW_SYMBOL = "Symbol"
COL_ENTRY_DATE = "Entry Date"
COL_COST_BASIS = "Cost Basis"
COL_UNREALIZED = "Unrealized"

SHOW_COLUMNS = [
    COL_POSITION_NUM,
    COL_SHOW_SYMBOL,
    COL_ENTRY_DATE,
    COL_COST_BASIS,
    COL_REALIZED,
    COL_UNREALIZED,
]

# --xlsx workbook layout. The "Positions" sheet is the --show rollup plus a "Closed" marker column;
# the "Legs" sheet is the per-leg analysis minus the Greeks.
COL_CLOSED = "Closed"

XLSX_POSITION_COLUMNS = [
    COL_POSITION_NUM,
    COL_SHOW_SYMBOL,
    COL_ENTRY_DATE,
    COL_COST_BASIS,
    COL_REALIZED,
    COL_UNREALIZED,
    COL_CLOSED,
]

XLSX_LEG_COLUMNS = [
    COL_CONTRACT,
    COL_POSITION_NUM,
    COL_POSITION_TYPE,
    COL_QUANTITY,
    COL_HELD,
    COL_TRADE_PRICE,
    COL_CUR_EXIT_PRICE,
    COL_REALIZED,
]

# Columns whose TEXT is colored by sign (green positive / red negative / black zero), per sheet.
XLSX_SIGN_FONT_COLUMNS = {
    "Positions": [COL_COST_BASIS],
    "Legs": [COL_QUANTITY, COL_HELD],
}
# Columns whose BACKGROUND is colored by sign (green positive / red negative / no fill at zero).
XLSX_SIGN_FILL_COLUMNS = {
    "Positions": [COL_REALIZED, COL_UNREALIZED],
    "Legs": [COL_REALIZED],
}

# Excel's conventional green/red pairing, so the sheets look native in Excel and Google Sheets alike.
_XL_GREEN_FONT = Font(color="FF006100")
_XL_RED_FONT = Font(color="FF9C0006")
_XL_GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
_XL_RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_XL_HEADER_FONT = Font(bold=True)

_XL_MONEY_FORMAT = "0.00"
_XL_INT_FORMAT = "0"
# Per-column Excel number formats; anything unlisted is left as text/general.
XLSX_NUMBER_FORMATS = {
    COL_POSITION_NUM: _XL_INT_FORMAT,
    COL_QUANTITY: _XL_INT_FORMAT,
    COL_HELD: _XL_INT_FORMAT,
    COL_TRADE_PRICE: _XL_MONEY_FORMAT,
    COL_CUR_EXIT_PRICE: _XL_MONEY_FORMAT,
    COL_COST_BASIS: _XL_MONEY_FORMAT,
    COL_REALIZED: _XL_MONEY_FORMAT,
    COL_UNREALIZED: _XL_MONEY_FORMAT,
}

_logger = getLogger(__name__)


def _is_int(text: str) -> bool:
    """True if text parses cleanly as an integer (e.g. '6', '-4' -- but not '', '1.0', 'x')."""
    try:
        int(text)
        return True
    except (TypeError, ValueError):
        return False


def _is_float(text: str) -> bool:
    """True if text parses as a float (e.g. '2.51', '0', '-1.5' -- but not '', 'x')."""
    try:
        float(text)
        return True
    except (TypeError, ValueError):
        return False


def _is_ib_datetime(text: str) -> bool:
    """True if text is an IB-style datetime, e.g. '20260717 09:58:57 US/Eastern' (or a bare date)."""
    try:
        get_datetime(text)
        return True
    except Exception:
        return False


def _is_symbol(text: str) -> bool:
    """True if text is a valid stock/option symbol, e.g. 'SPY' or 'SPY-C-20260821-800.0'."""
    try:
        SecurityDescriptor.from_string(text)
        return True
    except CoreException:
        return False


def _cell(row: pd.Series, col: str) -> str:
    """Returns a row's raw cell value as a stripped string ('' if missing or blank)."""
    value = row.get(col, "")
    return "" if value is None else str(value).strip()


def _row_problems(row: pd.Series, valid_types: set) -> List[str]:
    """
    Returns a list of type/format problems for a single CSV row (empty if the row is well-formed).

    Expected types: Position #, Quantity, Quantity Out are ints; Trade Price and Exit Price are
    floats; Date In is an IB-style datetime; Date Out is blank or an IB-style datetime; Position
    Type is one of POSITION_TYPE_MAP's descriptions; Symbol parses via SecurityDescriptor.
    """
    issues: List[str] = []

    if not _is_int(_cell(row, CSV_POSITION_NUM)):
        issues.append(f"Position # {_cell(row, CSV_POSITION_NUM)!r} is not an integer")

    date_in = _cell(row, CSV_DATE_IN)
    if not date_in:
        issues.append("Date In is blank")
    elif not _is_ib_datetime(date_in):
        issues.append(f"Date In {date_in!r} is not an IB-style datetime")

    ptype = _cell(row, CSV_POSITION_TYPE)
    if ptype not in valid_types:
        issues.append(f"Position Type {ptype!r} is not a recognized position type")

    if not _is_symbol(_cell(row, CSV_SYMBOL)):
        issues.append(f"Symbol {_cell(row, CSV_SYMBOL)!r} is not a valid symbol")

    if not _is_int(_cell(row, CSV_QUANTITY)):
        issues.append(f"Quantity {_cell(row, CSV_QUANTITY)!r} is not an integer")

    if not _is_float(_cell(row, CSV_TRADE_PRICE)):
        issues.append(f"Trade Price {_cell(row, CSV_TRADE_PRICE)!r} is not a number")

    date_out = _cell(row, CSV_DATE_OUT)
    if date_out and not _is_ib_datetime(date_out):
        issues.append(f"Date Out {date_out!r} is not an IB-style datetime or blank")

    if not _is_int(_cell(row, CSV_QUANTITY_OUT)):
        issues.append(f"Quantity Out {_cell(row, CSV_QUANTITY_OUT)!r} is not an integer")

    if not _is_float(_cell(row, CSV_EXIT_PRICE)):
        issues.append(f"Exit Price {_cell(row, CSV_EXIT_PRICE)!r} is not a number")

    return issues


def validate_positions_file(positions_file: str) -> List[str]:
    """
    Verifies the positions CSV is well-formed before any analysis runs.

    Confirms every column in REQUIRED_COLUMNS is present at the top of the file, then checks that
    every data row holds the expected type of data in each column (see _row_problems). Cells are read
    as raw strings so wrong types are caught rather than silently coerced by pandas.

    :param positions_file: path to the CSV
    :return: list of human-readable problem messages (each malformed row includes its full contents);
             empty if the file is valid
    """
    try:
        df = pd.read_csv(positions_file, dtype=str, keep_default_na=False)
    except Exception as ex:
        return [f"Could not read CSV file {positions_file!r}: {ex}"]

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        # Without the required headers there's no point checking row contents.
        return ["Missing required column(s): " + ", ".join(f"'{c}'" for c in missing)]

    valid_types = set(POSITION_TYPE_MAP.values())
    problems: List[str] = []
    for idx, row in df.iterrows():
        row_problems = _row_problems(row, valid_types)
        if row_problems:
            # +2: pandas rows are 0-based and the CSV's first line is the header, so row 0 is line 2.
            line_no = idx + 2
            contents = ", ".join(f"{col}={row[col]!r}" for col in REQUIRED_COLUMNS)
            problems.append(f"Row {line_no}: " + "; ".join(row_problems) + f"\n    {contents}")
    return problems


def load_positions(
    positions_file: str,
    symbol: Optional[str],
    expiration: Optional[str],
    position_num: Optional[int] = None,
    position_type: Optional[str] = None,
) -> pd.DataFrame:
    """
    Loads open positions from a CSV, optionally narrowing the legs by several filters.

    :param positions_file: path to CSV of open positions
    :param symbol: if given, keep only legs whose underlying matches (e.g. "SPY")
    :param expiration: if given, keep only legs with this IB-style expiration (e.g. "20260821")
    :param position_num: if given, keep only legs with this "Position #" value
    :param position_type: if given, keep only legs with this full "Position Type" (e.g. "Iron Condor")
    :return: filtered DataFrame of position rows (original CSV columns)
    """
    df = pd.read_csv(positions_file)

    # Parse each leg's descriptor so we can filter on underlying / expiration.
    descriptors = df[CSV_SYMBOL].apply(SecurityDescriptor)
    df = df.assign(
        _ticker=descriptors.apply(lambda d: d.ticker),
        _expiration=descriptors.apply(lambda d: d.expiration),
    )

    if symbol:
        df = df[df["_ticker"] == symbol]
    if expiration:
        df = df[df["_expiration"] == expiration]
    if position_num is not None:
        df = df[df[CSV_POSITION_NUM] == position_num]
    if position_type:
        df = df[df[CSV_POSITION_TYPE] == position_type]

    return df.reset_index(drop=True)


async def collect_leg_data(
    option_manager: OptionDataManager, positions_df: pd.DataFrame
) -> Tuple[OptionData, List[Optional[OptionInfo]]]:
    """
    Fetches current market data (price + Greeks) for every leg in positions_df.

    Uses OptionDataManager to retrieve an OptionInfo per contract and collects them into an
    OptionData object so the results are available as a pandas DataFrame.

    :param option_manager: connected OptionDataManager
    :param positions_df: filtered positions
    :return: (OptionData holding one row per successfully-fetched leg, per-row OptionInfo list
              aligned with positions_df -- None where a fetch failed)
    """
    option_data = OptionData("positions", current_datetime())
    infos: List[Optional[OptionInfo]] = []

    for _, row in positions_df.iterrows():
        descriptor = SecurityDescriptor(row[CSV_SYMBOL])
        try:
            option_info = await option_manager.get_option_info(
                ticker=descriptor.ticker,
                expiration=descriptor.expiration,
                right=descriptor.right,
                strike=descriptor.strike,
            )
        except OptionDataException as ex:
            print(f"WARNING: could not get data for {row[CSV_SYMBOL]}: {ex}")
            option_info = None

        infos.append(option_info)
        if option_info is not None:
            option_data.add_data(option_info)

    return option_data, infos


def _csv_int(pos_row: pd.Series, col: str) -> int:
    """Reads an integer CSV cell, treating a blank/missing/NaN value as 0."""
    value = pos_row.get(col, 0)
    if value == "" or value is None or pd.isna(value):
        return 0
    return int(value)


def _csv_float(pos_row: pd.Series, col: str) -> float:
    """Reads a float CSV cell, treating a blank/missing/NaN value as 0.0."""
    value = pos_row.get(col, 0.0)
    if value == "" or value is None or pd.isna(value):
        return 0.0
    return float(value)


def build_output_dataframe(positions_df: pd.DataFrame, infos: List[Optional[OptionInfo]]) -> pd.DataFrame:
    """
    Combines the CSV position details with fetched per-contract market data into one numeric
    DataFrame -- one row per leg -- ready for aggregation and display.

    Two derived per-leg quantities are added:
      * "Held": the number of contracts ACTUALLY held now, = Quantity + Quantity Out (signed; can be 0 once a
        leg is fully closed). All market-value/Greek aggregation uses this, not the original entry Quantity.
      * "Realized": dollars of realized P/L on the closed portion, = -Quantity Out * (Exit Price - Trade Price)
        * contract multiplier. This is sign-correct for both long and short legs and is 0 until an exit is
        recorded.

    "Cur/Exit Price" is the current market price while a leg still holds contracts, and the CSV's average Exit
    Price once it is flat (Held == 0).

    :param positions_df: filtered positions (CSV columns)
    :param infos: OptionInfo per leg (aligned with positions_df), None where unavailable
    :return: DataFrame with OUTPUT_COLUMNS
    """
    rows = []
    for (_, pos_row), info in zip(positions_df.iterrows(), infos):
        quantity = _csv_int(pos_row, CSV_QUANTITY)
        quantity_out = _csv_int(pos_row, CSV_QUANTITY_OUT)
        trade_price = _csv_float(pos_row, CSV_TRADE_PRICE)
        exit_price = _csv_float(pos_row, CSV_EXIT_PRICE)

        held = quantity + quantity_out
        realized = -quantity_out * (exit_price - trade_price) * CONTRACT_MULTIPLIER

        # A flat leg holds nothing, so a live quote says nothing about it; report the average price it was
        # closed at instead. Either way the leg contributes price * held = 0 to the aggregate, so this is a
        # display choice only.
        cur_exit_price = exit_price if held == 0 else (info.price if info else float("nan"))

        rows.append(
            {
                COL_CONTRACT: pos_row[CSV_SYMBOL],
                COL_POSITION_NUM: pos_row[CSV_POSITION_NUM],
                COL_POSITION_TYPE: pos_row[CSV_POSITION_TYPE],
                COL_QUANTITY: quantity,
                COL_HELD: held,
                COL_TRADE_PRICE: trade_price,
                COL_CUR_EXIT_PRICE: cur_exit_price,
                COL_REALIZED: realized,
                COL_IV: info.implied_volatility if info else float("nan"),
                COL_DELTA: info.delta if info else float("nan"),
                COL_THETA: info.theta if info else float("nan"),
                COL_GAMMA: info.gamma if info else float("nan"),
                COL_VEGA: info.vega if info else float("nan"),
            }
        )
    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


def build_aggregate_row(df: pd.DataFrame) -> Tuple[dict, float, float]:
    """
    Computes the aggregate row for a set of option legs.

    All market-value and Greek aggregation is done on the contracts ACTUALLY HELD ("Held" = Quantity +
    Quantity Out), so fully-closed legs (Held == 0) drop out naturally.

    Standard rules for combining option positions:
      * Position Greek = sum over legs of (per-contract Greek * held * contract multiplier).
        Held is negative for short legs, so shorts subtract as expected.
      * Aggregate trade/current "price" is expressed as net dollars: sum(price * held * multiplier).
        A positive value is a net debit (cash paid); a negative value is a net credit (cash
        received). Flat legs carry the exit price rather than a live quote, but held == 0 zeroes the
        term either way, so the aggregate is still purely the value of what is held.
      * Realized P/L is summed straight from the per-leg "Realized" dollars (closed portion).
      * Implied volatility does not aggregate meaningfully across strikes, so it is left blank.

    :param df: per-leg output DataFrame (OUTPUT_COLUMNS)
    :return: (aggregate row dict keyed by OUTPUT_COLUMNS, unrealized P/L in dollars, realized P/L in dollars)
    """
    held = df[COL_HELD]
    mult = CONTRACT_MULTIPLIER

    net_trade = (df[COL_TRADE_PRICE] * held * mult).sum()
    net_current = (df[COL_CUR_EXIT_PRICE] * held * mult).sum()
    unrealized_pl = net_current - net_trade
    realized_pl = df[COL_REALIZED].sum()

    aggregate = {
        COL_CONTRACT: "AGGREGATE",
        COL_POSITION_NUM: "",
        COL_POSITION_TYPE: "",
        COL_QUANTITY: df[COL_QUANTITY].sum(),
        COL_HELD: held.sum(),
        COL_TRADE_PRICE: net_trade,
        COL_CUR_EXIT_PRICE: net_current,
        COL_REALIZED: realized_pl,
        COL_IV: float("nan"),
        COL_DELTA: (df[COL_DELTA] * held * mult).sum(),
        COL_THETA: (df[COL_THETA] * held * mult).sum(),
        COL_GAMMA: (df[COL_GAMMA] * held * mult).sum(),
        COL_VEGA: (df[COL_VEGA] * held * mult).sum(),
    }
    return aggregate, unrealized_pl, realized_pl


def _fmt(value, decimals: int) -> str:
    """Formats a numeric cell, rendering NaN/empty as a dash."""
    if value == "" or value is None:
        return ""
    try:
        if pd.isna(value):
            return "-"
    except (TypeError, ValueError):
        return str(value)
    if value == 0:  # normalize -0.0 so it prints as 0.00, not -0.00
        value = 0.0
    return f"{value:.{decimals}f}"


def format_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Returns a copy of df with each column formatted to a readable string."""
    decimals = {
        COL_TRADE_PRICE: 2,
        COL_CUR_EXIT_PRICE: 2,
        COL_REALIZED: 2,
        COL_IV: 4,
        COL_DELTA: 4,
        COL_THETA: 4,
        COL_GAMMA: 5,
        COL_VEGA: 4,
    }
    display = pd.DataFrame(columns=OUTPUT_COLUMNS)
    for col in OUTPUT_COLUMNS:
        if col in decimals:
            display[col] = df[col].apply(lambda v, d=decimals[col]: _fmt(v, d))
        elif col in (COL_QUANTITY, COL_HELD):
            display[col] = df[col].apply(lambda v: "" if v == "" else str(int(v)))
        else:
            display[col] = df[col].astype(str)
    return display


def print_analysis(df: pd.DataFrame, aggregate: dict, unrealized_pl: float, realized_pl: float):
    """Pretty-prints the per-leg table, the aggregate row, and a summary."""
    # Format legs and aggregate together so the columns align in a single table.
    combined = pd.concat([df, pd.DataFrame([aggregate], columns=OUTPUT_COLUMNS)], ignore_index=True)
    display = format_for_display(combined)
    lines = display.to_string(index=False).splitlines()
    table_width = max(len(line) for line in lines)

    print("\nPer-leg analysis")
    print("=" * table_width)
    # All rows except the last are individual legs; the last row is the aggregate.
    for line in lines[:-1]:
        print(line)
    print("-" * table_width)
    print(lines[-1])
    print("=" * table_width)

    print("\nPosition summary")
    print("-" * 40)
    print(f"  Net premium (debit +/credit -): {aggregate[COL_TRADE_PRICE]:>12.2f}")
    print(f"  Current net value             : {aggregate[COL_CUR_EXIT_PRICE]:>12.2f}")
    print(f"  Unrealized P/L                : {unrealized_pl:>12.2f}")
    print(f"  Realized P/L                  : {realized_pl:>12.2f}")
    print(f"  Total P/L (real + unreal)     : {realized_pl + unrealized_pl:>12.2f}")
    print(f"  Position delta                : {aggregate[COL_DELTA]:>12.4f}")
    print(f"  Position theta                : {aggregate[COL_THETA]:>12.4f}")
    print(f"  Position gamma                : {aggregate[COL_GAMMA]:>12.4f}")
    print(f"  Position vega                 : {aggregate[COL_VEGA]:>12.4f}")
    print()
    print("Note: 'Held' = contracts actually held now (Quantity + Quantity Out); market value and Greeks")
    print("      aggregate over Held. 'Cur/Exit Price' is the current price while contracts are held, or the")
    print("      average exit price from the CSV once the leg is flat (Held = 0). Aggregate Trade/Cur-Exit")
    print("      Price are net dollars (price * held * 100); aggregate Greeks are position Greeks")
    print("      (per-contract Greek * held * 100). Per-leg 'Realized' is closed-portion P/L in dollars.")
    print("      Quantities: '-' = short.")
    print()


def _entry_date_token(value) -> str:
    """Extracts the date portion (e.g. '20260713') from a 'Date In' cell, or '' if blank/missing."""
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    # 'Date In' looks like '20260713 09:30:00 US/Eastern' (or just '20260716'); keep the date token.
    return text.split()[0]


def build_show_dataframe(
    positions_df: pd.DataFrame, infos: List[Optional[OptionInfo]], closed: bool = False
) -> pd.DataFrame:
    """
    Rolls the per-leg CSV rows up into one summary row per whole position for the --show tables.

    A position counts as closed once every one of its legs is flat (Held == 0 for all legs), and as held
    while any leg still has contracts. `closed` selects which of the two sets to return, so the same
    rollup feeds both the "Positions held" and "Closed positions" tables.

    Per position the following are computed:
      * Symbol: the underlying ticker the legs share (joined with '/' in the unusual case they differ).
      * Entry Date: the earliest 'Date In' date across the legs.
      * Cost Basis: net entry premium = sum(Trade Price * Quantity * 100). Positive is a net debit
        paid; negative is a net credit collected.
      * Realized: dollars of realized P/L on any closed portions = sum(-Quantity Out * (Exit Price -
        Trade Price) * 100).
      * Unrealized: dollars of unrealized P/L on contracts still held = sum((Current Price -
        Trade Price) * Held * 100).

    For a closed position every leg is flat, so its Unrealized is always 0.00 -- the column is kept anyway
    so both tables share one layout.

    :param positions_df: positions loaded from the CSV (one row per leg)
    :param infos: OptionInfo per leg (aligned with positions_df), None where market data was unavailable
    :param closed: True to return fully-closed positions; False (default) for positions still held
    :return: DataFrame with SHOW_COLUMNS, one row per matching position
    """
    records = []
    for (_, pos_row), info in zip(positions_df.iterrows(), infos):
        descriptor = SecurityDescriptor(pos_row[CSV_SYMBOL])
        quantity = _csv_int(pos_row, CSV_QUANTITY)
        quantity_out = _csv_int(pos_row, CSV_QUANTITY_OUT)
        trade_price = _csv_float(pos_row, CSV_TRADE_PRICE)
        exit_price = _csv_float(pos_row, CSV_EXIT_PRICE)

        held = quantity + quantity_out
        realized = -quantity_out * (exit_price - trade_price) * CONTRACT_MULTIPLIER
        cost_basis = trade_price * quantity * CONTRACT_MULTIPLIER
        # A fully-closed leg (held == 0) contributes no unrealized P/L, and needs no market data.
        if held != 0 and info is not None:
            unrealized = (info.price - trade_price) * held * CONTRACT_MULTIPLIER
        else:
            unrealized = 0.0

        records.append(
            {
                "pos_num": pos_row[CSV_POSITION_NUM],
                "ticker": descriptor.ticker,
                "entry": _entry_date_token(pos_row.get(CSV_DATE_IN)),
                "held_abs": abs(held),
                "cost_basis": cost_basis,
                "realized": realized,
                "unrealized": unrealized,
            }
        )

    legs = pd.DataFrame(records)
    rows = []
    # sort=False keeps positions in the order they first appear in the CSV.
    for pos_num, group in legs.groupby("pos_num", sort=False):
        # Nothing left held on any leg == fully closed; keep whichever set the caller asked for.
        if (group["held_abs"].sum() == 0) != closed:
            continue
        tickers = sorted(set(group["ticker"]))
        entries = [e for e in group["entry"] if e]
        rows.append(
            {
                COL_POSITION_NUM: pos_num,
                COL_SHOW_SYMBOL: "/".join(tickers),
                COL_ENTRY_DATE: min(entries) if entries else "",
                COL_COST_BASIS: group["cost_basis"].sum(),
                COL_REALIZED: group["realized"].sum(),
                COL_UNREALIZED: group["unrealized"].sum(),
            }
        )
    return pd.DataFrame(rows, columns=SHOW_COLUMNS)


def print_show_table(df: pd.DataFrame, title: str):
    """
    Pretty-prints one --show summary table (one row per position) under `title`, with a totals row.

    :param df: summary rows with SHOW_COLUMNS (assumed non-empty; empty tables are skipped by the caller)
    :param title: heading for the table, e.g. "Positions held" or "Closed positions"
    """
    total = {
        COL_POSITION_NUM: "TOTAL",
        COL_SHOW_SYMBOL: "",
        COL_ENTRY_DATE: "",
        COL_COST_BASIS: df[COL_COST_BASIS].sum(),
        COL_REALIZED: df[COL_REALIZED].sum(),
        COL_UNREALIZED: df[COL_UNREALIZED].sum(),
    }
    combined = pd.concat([df, pd.DataFrame([total], columns=SHOW_COLUMNS)], ignore_index=True)

    display = pd.DataFrame(columns=SHOW_COLUMNS)
    display[COL_POSITION_NUM] = combined[COL_POSITION_NUM].astype(str)
    display[COL_SHOW_SYMBOL] = combined[COL_SHOW_SYMBOL].astype(str)
    display[COL_ENTRY_DATE] = combined[COL_ENTRY_DATE].astype(str)
    for col in (COL_COST_BASIS, COL_REALIZED, COL_UNREALIZED):
        display[col] = combined[col].apply(lambda v: _fmt(v, 2))

    lines = display.to_string(index=False).splitlines()
    table_width = max(len(line) for line in lines)

    print(f"\n{title}")
    print("=" * table_width)
    # All rows except the last are individual positions; the last row is the totals row.
    for line in lines[:-1]:
        print(line)
    print("-" * table_width)
    print(lines[-1])
    print("=" * table_width)


def print_show_note():
    """Prints the footnote shared by the --show tables (once, after whichever tables were displayed)."""
    print("\nNote: Cost Basis is net entry premium (Trade Price * Quantity * 100); negative = credit")
    print("      collected. Realized is closed-portion P/L; Unrealized is on contracts still held,")
    print("      and is therefore always 0.00 for closed positions.")
    print()


def build_positions_page_xlsx(positions_df: pd.DataFrame, infos: List[Optional[OptionInfo]]) -> pd.DataFrame:
    """
    Builds the "Positions" sheet for --xlsx output. This function serves XLSX generation only; the --show
    tables call build_show_dataframe() directly, since they print held and closed positions separately.

    Uses the same rollup as those tables, but merges held and closed positions into one frame ordered by
    Position #, with a "Closed" column marking the closed ones with "X".

    :param positions_df: positions loaded from the CSV (one row per leg)
    :param infos: OptionInfo per leg (aligned with positions_df), None where market data was unavailable
    :return: DataFrame with XLSX_POSITION_COLUMNS
    """
    held = build_show_dataframe(positions_df, infos)
    closed = build_show_dataframe(positions_df, infos, closed=True)
    held[COL_CLOSED] = ""
    closed[COL_CLOSED] = "X"

    combined = pd.concat([held, closed], ignore_index=True)
    if len(combined) == 0:
        return pd.DataFrame(columns=XLSX_POSITION_COLUMNS)
    # Position # comes from the CSV as text in some rows, so sort on a numeric view of it.
    order = pd.to_numeric(combined[COL_POSITION_NUM], errors="coerce")
    combined = combined.assign(_order=order).sort_values("_order", kind="stable").drop(columns="_order")
    return combined.reset_index(drop=True)[XLSX_POSITION_COLUMNS]


def _xl_number(value):
    """Converts a cell value for Excel: NaN/None become blank cells, everything else passes through."""
    try:
        if value is None or pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _xl_sign(value) -> int:
    """Returns 1, -1, or 0 for a cell's sign; blanks and non-numerics count as 0 (no coloring)."""
    try:
        if value is None or pd.isna(value):
            return 0
        number = float(value)
    except (TypeError, ValueError):
        return 0
    if number > 0:
        return 1
    return -1 if number < 0 else 0


def _write_xlsx_sheet(worksheet, df: pd.DataFrame, columns: List[str], font_cols: List[str], fill_cols: List[str]):
    """
    Writes one DataFrame to a worksheet: bold frozen header, per-column number formats, sign coloring, and
    column widths sized to the content.

    :param worksheet: openpyxl worksheet to fill
    :param df: rows to write
    :param columns: columns to write, in order
    :param font_cols: columns whose text is colored by sign
    :param fill_cols: columns whose background is colored by sign
    """
    worksheet.append(list(columns))
    for cell in worksheet[1]:
        cell.font = _XL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        worksheet.append([_xl_number(row[col]) for col in columns])

    for col_index, col in enumerate(columns, start=1):
        number_format = XLSX_NUMBER_FORMATS.get(col)
        color_font = col in font_cols
        color_fill = col in fill_cols
        widest = len(str(col))

        for row_index, value in enumerate(df[col], start=2):
            cell = worksheet.cell(row=row_index, column=col_index)
            if number_format:
                cell.number_format = number_format
            sign = _xl_sign(value) if (color_font or color_fill) else 0
            if color_font and sign:
                cell.font = _XL_GREEN_FONT if sign > 0 else _XL_RED_FONT
            if color_fill and sign:
                cell.fill = _XL_GREEN_FILL if sign > 0 else _XL_RED_FILL
            # Approximate the rendered width: money columns show two decimals.
            if number_format == _XL_MONEY_FORMAT and _xl_number(value) is not None:
                shown = f"{float(value):.2f}"
            else:
                shown = "" if _xl_number(value) is None else str(value)
            widest = max(widest, len(shown))

        worksheet.column_dimensions[get_column_letter(col_index)].width = widest + 3

    worksheet.freeze_panes = "A2"


def write_xlsx(path: str, positions_page: pd.DataFrame, legs_page: pd.DataFrame):
    """
    Writes the two-sheet workbook: "Positions" (one row per position) and "Legs" (one row per CSV leg).

    Coloring follows the sign of the value -- green for positive, red for negative, untouched at zero.
    "Qty", "Held", and "Cost Basis" color their text; "Realized" and "Unrealized" color the cell
    background.

    :param path: output .xlsx path
    :param positions_page: rows with XLSX_POSITION_COLUMNS
    :param legs_page: rows with XLSX_LEG_COLUMNS
    """
    workbook = Workbook()
    positions_sheet = workbook.active
    positions_sheet.title = "Positions"
    legs_sheet = workbook.create_sheet("Legs")

    _write_xlsx_sheet(
        positions_sheet,
        positions_page,
        XLSX_POSITION_COLUMNS,
        XLSX_SIGN_FONT_COLUMNS["Positions"],
        XLSX_SIGN_FILL_COLUMNS["Positions"],
    )
    _write_xlsx_sheet(
        legs_sheet,
        legs_page,
        XLSX_LEG_COLUMNS,
        XLSX_SIGN_FONT_COLUMNS["Legs"],
        XLSX_SIGN_FILL_COLUMNS["Legs"],
    )
    workbook.save(path)


def calculate_expected_move_loss(delta: float, gamma: float, expected_move: float) -> float:
    """
    Estimates a position's one-day loss if the underlying makes its expected (one standard deviation) move
    against it, using the position's delta/gamma exposure:

        expected move loss = abs(delta) * move + 0.5 * abs(gamma) * move^2

    Both Greeks are taken in absolute value, so the result is the loss under whichever direction hurts: the
    delta term is the directional exposure, and the gamma term is the second-order add-on that a short-gamma
    position pays on a large move.

    This is a local estimate at one standard deviation, NOT the position's maximum possible loss -- the true
    worst case is set by the position's structure at expiration and is generally much larger.

    :param delta: position delta (aggregated over all legs, per $1 of underlying)
    :param gamma: position gamma (aggregated over all legs)
    :param expected_move: one standard deviation move in the underlying, in dollars
    :return: estimated expected-move loss in dollars (always non-negative)
    """
    return abs(delta) * expected_move + 0.5 * abs(gamma) * expected_move * expected_move


async def print_expected_moves(data_driver: BaseDriver, tickers: List[str], delta: float, gamma: float):
    """
    Prints each underlying's expected 1-day move (one standard deviation), using the broker's current implied
    volatility and most recent price: expected move = price * IV * sqrt(1 / 365) (see calculate_expected_move).
    Each expected move is then turned into an expected move loss for the position as a whole (see
    calculate_expected_move_loss).

    :param data_driver: connected broker driver
    :param tickers: distinct underlying symbols to report on
    :param delta: position delta, aggregated over all filtered legs
    :param gamma: position gamma, aggregated over all filtered legs
    """
    print("\nExpected next-day move (1 std dev)")
    print("-" * 60)
    moves = {}
    for ticker in tickers:
        implied_vol = await data_driver.get_implied_volatility(ticker)
        recent, error_str = await data_driver.get_most_recent_data(
            ticker, bar_size=BarSize.ONE_DAY, request_info_type=RequestedInfoType.TRADES
        )
        if implied_vol is None or implied_vol <= 0.0 or recent is None:
            print(f"  {ticker:<6} expected move unavailable (implied volatility or price missing)")
            continue
        price = recent[0]["close"]
        move = calculate_expected_move(price, implied_vol, 1)
        moves[ticker] = move
        print(
            f"  {ticker:<6} price {price:>10.2f}  IV {implied_vol:>7.4f}  "
            f"move +/-{move:>7.2f}  ({price - move:.2f} to {price + move:.2f})"
        )
    print()

    if not moves:
        return

    print("Expected move loss")
    print("-" * 60)
    print(f"  Position delta {delta:>10.4f}   gamma {gamma:>10.5f}")
    for ticker, move in moves.items():
        move_loss = calculate_expected_move_loss(delta, gamma, move)
        print(f"  {ticker:<6} move +/-{move:>7.2f}  expected move loss {move_loss:>12.2f}")
    print()
    print("Note: expected move loss = |delta| * move + 0.5 * |gamma| * move^2, using the position's")
    print("      aggregate delta and gamma against a one standard deviation move in the underlying.")
    print("      This is the loss if the underlying moves one std dev against the position; it is NOT")
    print("      the position's maximum possible loss, which is set by its structure at expiration.")
    if len(moves) > 1:
        print("      The position spans more than one underlying, so the aggregate Greeks cover all of")
        print("      them; each row moves only its own underlying. Filter with --symbol for a per-")
        print("      underlying figure.")
    print()


async def main(parser: argparse.ArgumentParser):
    """Top-level function: unpacks arguments, fetches data, and prints the analysis."""
    args = parser.parse_args()

    basicConfig(filename="position_analyzer.log", level=INFO)

    # Verify the CSV is well-formed as soon as it's loaded. If anything is wrong, report every
    # problem (with the offending rows) and exit gracefully without doing anything else.
    problems = validate_positions_file(args.positions_file)
    if problems:
        print(f"CSV validation failed for {args.positions_file}. The tool will not run until these are fixed:\n")
        for problem in problems:
            print(f"  - {problem}")
        print("\nRecognized position types: " + ", ".join(sorted(set(POSITION_TYPE_MAP.values()))))
        return

    # A broker must be chosen explicitly; there's no sensible default.
    if not args.ib and not args.schwab:
        print("No broker specified. Pass --ib (Interactive Brokers) or --schwab (Schwab).")
        return

    # Resolve the short --position-type code (e.g. "IC") to the full CSV name (e.g. "Iron Condor").
    position_type = POSITION_TYPE_MAP[args.position_type] if args.position_type else None

    # --show and --xlsx both cover the whole CSV (held and closed alike), so they ignore the narrowing
    # filters; only the per-leg analysis honors them.
    whole_csv = args.show or args.xlsx
    if whole_csv:
        positions_df = load_positions(args.positions_file, None, None)
    else:
        positions_df = load_positions(
            args.positions_file,
            args.symbol,
            args.expiration,
            position_num=args.position_num,
            position_type=position_type,
        )
    if len(positions_df) == 0:
        print("No positions found." if whole_csv else "No positions match the given filters.")
        return

    if whole_csv:
        print(f"Showing positions from {args.positions_file}")
    else:
        print(f"Analyzing {len(positions_df)} option leg(s) from {args.positions_file}")
        if args.symbol:
            print(f"  Filtered to underlying: {args.symbol}")
        if args.expiration:
            print(f"  Filtered to expiration: {args.expiration}")
        if args.position_num is not None:
            print(f"  Filtered to position #: {args.position_num}")
        if position_type:
            print(f"  Filtered to position type: {position_type} ({args.position_type})")

    print("Please wait...")
    option_manager = OptionDataManager()
    if args.schwab:
        data_driver = SchwabDriver.create()
        connect_hint = "Check your Schwab credentials in .env and the token file, then try again."
    else:
        data_driver = IBDriver.create(sim_account=True, client_id=CLIENT_ID)
        connect_hint = "Make sure IB Gateway or TWS is running and logged in to the paper/sim account, then try again."
    option_manager.add_driver(data_driver)

    if not data_driver.is_connected():
        print(f"Could not connect to the broker. {connect_hint}")
        data_driver.disconnect()
        return

    try:
        _, infos = await collect_leg_data(option_manager, positions_df)

        if args.show:
            held_df = build_show_dataframe(positions_df, infos)
            closed_df = build_show_dataframe(positions_df, infos, closed=True)
            # Each table is skipped when it has no rows; the shared note follows whichever were shown.
            if len(held_df) == 0 and len(closed_df) == 0:
                print("No positions found.")
            else:
                if len(held_df) > 0:
                    print_show_table(held_df, "Positions held")
                if len(closed_df) > 0:
                    print_show_table(closed_df, "Closed positions")
                print_show_note()

        if args.xlsx:
            positions_page = build_positions_page_xlsx(positions_df, infos)
            legs_page = build_output_dataframe(positions_df, infos)[XLSX_LEG_COLUMNS]
            write_xlsx(args.xlsx, positions_page, legs_page)
            print(f"\nWrote {len(positions_page)} position(s) and {len(legs_page)} leg(s) to {args.xlsx}")

        if whole_csv:
            return

        output_df = build_output_dataframe(positions_df, infos)
        aggregate, unrealized_pl, realized_pl = build_aggregate_row(output_df)
        print_analysis(output_df, aggregate, unrealized_pl, realized_pl)

        # Expected next-day move for each distinct underlying among the filtered legs.
        tickers = sorted({SecurityDescriptor(symbol).ticker for symbol in positions_df[CSV_SYMBOL]})
        await print_expected_moves(data_driver, tickers, aggregate[COL_DELTA], aggregate[COL_GAMMA])
    except asyncio.CancelledError:
        print("Program cancelled by user.")
    except Exception as ex:
        print(f"Got exception: {ex}")
        print(traceback.format_exc())
    finally:
        data_driver.disconnect()


def build_parser() -> argparse.ArgumentParser:
    """Builds the command-line argument parser."""
    parser = argparse.ArgumentParser(
        prog="python -m scripts.position_analyzer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent("""\
            Analyze a set of open options positions read from a CSV.

            For each option leg the tool reports the current per-contract price, implied volatility,
            and Greeks (delta, theta, gamma, vega) next to the trade details from the CSV, along with
            the contracts actually held (Quantity + Quantity Out) and realized P/L on any closed
            portion. A final aggregate row summarizes the filtered legs as a whole: net position
            Greeks, net value, unrealized profit/loss, and realized profit/loss.

            A broker must be selected with --ib (Interactive Brokers; requires IB Gateway or TWS
            running locally, paper/sim account) or --schwab (Charles Schwab; requires credentials
            in .env).
            """),
        epilog=textwrap.dedent("""\
            Examples:
              # Analyze every open position in the CSV, using Interactive Brokers for market data
              python -m scripts.position_analyzer --positions-file .\\data\\options_trades_2026.csv --ib

              # Same, but using Schwab for market data
              python -m scripts.position_analyzer --positions-file .\\data\\options_trades_2026.csv --schwab

              # Narrow to a single underlying
              python -m scripts.position_analyzer --positions-file .\\data\\options_trades_2026.csv --ib --symbol SPY

              # Narrow to a single underlying and expiration
              python -m scripts.position_analyzer --positions-file .\\data\\options_trades_2026.csv --ib --symbol QQQ --expiration 20260821

              # Narrow to one position number, or to all positions of a given type
              python -m scripts.position_analyzer --positions-file .\\data\\options_trades_2026.csv --ib --position-num 2
              python -m scripts.position_analyzer --positions-file .\\data\\options_trades_2026.csv --ib --position-type IC

              # Write an Excel workbook ('Positions' and 'Legs' sheets) for import into Google Drive
              python -m scripts.position_analyzer --positions-file .\\data\\current_positions.csv --schwab --xlsx positions.xlsx

              # Show one-row-per-position summaries of everything held and everything closed
              # (current_positions.csv carries the Date In / Quantity Out / Exit Price columns that
              #  entry date and realized P/L are computed from)
              python -m scripts.position_analyzer --positions-file .\\data\\current_positions.csv --ib --show

            Notes:
              * The CSV must have columns: 'Position #', 'Date In', 'Position Type', 'Symbol',
                'Quantity', 'Trade Price', 'Date Out', 'Quantity Out', 'Exit Price' (the
                current_positions.csv format). Symbols are IB-style, e.g. SPY-C-20260821-800.0.
              * The CSV is validated up front: if any required column is missing or any row has a
                badly-typed field, the offending rows are printed and the tool exits without doing
                anything else.
              * A negative Quantity indicates a short (sold) leg. 'Quantity Out' accumulates the
                signed closing trades, so contracts still held = Quantity + Quantity Out.
            """),
    )
    parser.add_argument(
        "--positions-file",
        help="Path to a CSV containing open positions.",
        required=True,
        type=str,
    )
    broker_group = parser.add_mutually_exclusive_group()
    broker_group.add_argument(
        "--ib",
        help="Use Interactive Brokers for market data (requires IB Gateway or TWS running locally).",
        action="store_true",
    )
    broker_group.add_argument(
        "--schwab",
        help="Use Charles Schwab for market data (requires Schwab credentials in .env).",
        action="store_true",
    )
    parser.add_argument(
        "--symbol",
        help="Narrow analysis to options sharing this underlying, e.g. SPY or QQQ.",
        required=False,
        default=None,
        type=str,
    )
    parser.add_argument(
        "--expiration",
        help="Narrow analysis to options with this IB-style expiration, e.g. 20260821.",
        required=False,
        default=None,
        type=str,
    )
    parser.add_argument(
        "--position-num",
        help="Narrow analysis to legs with this 'Position #' from the CSV, e.g. 2.",
        required=False,
        default=None,
        type=int,
    )
    parser.add_argument(
        "--position-type",
        help="Narrow analysis to legs of this position type. "
        + ", ".join(f"{code}={name}" for code, name in POSITION_TYPE_MAP.items())
        + ".",
        required=False,
        default=None,
        choices=list(POSITION_TYPE_MAP.keys()),
        type=str,
    )
    parser.add_argument(
        "--show",
        help="Show summary tables of every position in the CSV -- 'Positions held' and 'Closed "
        "positions' -- one row per position (symbol, entry date, cost basis, realized and unrealized "
        "P/L). Either table is omitted when it has no positions. The narrowing filters are ignored.",
        action="store_true",
    )
    parser.add_argument(
        "--xlsx",
        help="Write an Excel workbook to this path, with a 'Positions' sheet (one row per position, "
        "closed ones marked 'X') and a 'Legs' sheet (one row per CSV leg). Values are colored by sign. "
        "Covers the whole CSV, so the narrowing filters are ignored.",
        type=str,
    )
    return parser


if __name__ == "__main__":
    asyncio.run(main(build_parser()))
